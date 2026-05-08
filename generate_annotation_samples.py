"""
generate_annotation_samples.py
--------------------------------
Generates 30 explanation samples with full text for the manual
annotation study (MC2 response).

Runs 6 samples × 5 strategies = 30 explanation calls using Qwen 2.5 7B.
Saves full explanation texts + symbolic evidence to:
  - annotation_samples.csv   (for computing agreement)
  - annotation_sheet_A.xlsx  (blank sheet for Annotator A)
  - annotation_sheet_B.xlsx  (blank sheet for Annotator B)

Run from safe_gem folder:
    python generate_annotation_samples.py
"""

import numpy as np
import os, sys, csv, random

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

DATA_DIR    = r"C:\Users\Valued User\Desktop\Office PhD\agentic_egs\agentic_egs\agentic_egs\Dataset"
RESULTS_DIR = r"C:\Users\Valued User\Desktop\Office PhD\agentic_egs\agentic_egs\agentic_egs\Results"

SAMPLING_RATE = 500.0
STRATEGIES    = ["P1_raw", "P2_grounded", "P3_uncertainty",
                 "P4_constrained", "P5_safe_gem"]
N_PER_STRAT   = 6   # 6 samples × 5 strategies = 30 explanations
RANDOM_SEED   = 42


def select_samples(y, pred, conf, seed=42):
    """Select 6 diverse samples for annotation."""
    rng  = random.Random(seed)
    idx  = list(range(len(y)))
    # 2 clear events, 2 wrong noise, 2 ambiguous
    ev   = [i for i in idx if y[i]==1 and pred[i]==1 and conf[i]>0.95][:2]
    no   = [i for i in idx if y[i]==0 and pred[i]==1 and conf[i]>0.95][:2]
    amb  = sorted([i for i in idx if 0.45 < conf[i] < 0.65],
                  key=lambda i: abs(conf[i]-0.5))[:2]
    return ev + no + amb


def ev_summary(ev):
    """Short string summary of symbolic evidence."""
    parts = []
    if ev.snr_level not in ("unavailable",""):
        parts.append(f"SNR:{ev.snr_level}")
    if ev.frequency_band not in ("unavailable",""):
        parts.append(f"Freq:{ev.frequency_band}")
    if ev.onset_character not in ("unavailable",""):
        parts.append(f"Onset:{ev.onset_character}")
    if ev.amplitude_contrast not in ("unavailable",""):
        parts.append(f"Amp:{ev.amplitude_contrast}")
    if ev.das_coherence not in ("unavailable",""):
        parts.append(f"DAS:{ev.das_coherence}")
    if ev.event_support not in ("unavailable",""):
        parts.append(f"EvtSupport:{ev.event_support}")
    if ev.signal_ambiguity not in ("unavailable",""):
        parts.append(f"Ambiguity:{ev.signal_ambiguity}")
    return " | ".join(parts)


if __name__ == "__main__":
    from safe_gem import SafeGEMPipeline
    from generate_explanation import ExplanationGenerator

    X    = np.load(os.path.join(DATA_DIR, "X_forge.npy"), mmap_mode='r')
    y    = np.load(os.path.join(DATA_DIR, "y_forge.npy"))
    pred = np.load(os.path.join(DATA_DIR, "forge_predictions.npy"))
    conf = np.load(os.path.join(DATA_DIR, "forge_confidences.npy"))

    sample_indices = select_samples(y, pred, conf)
    print(f"Selected {len(sample_indices)} samples: {sample_indices}")

    pipeline = SafeGEMPipeline(pre_event_s=0.5)
    gen      = ExplanationGenerator(model="qwen2.5:7b")
    gen.warmup()

    # Only use one strategy per sample for annotation variety
    # (6 samples × 5 strategies = one strategy per sample pair)
    rows = []
    call_num = 0

    for strat_idx, strategy in enumerate(STRATEGIES):
        sample_idx = sample_indices[strat_idx]   # pair strategy with sample
        idx = sample_idx

        sample_id = f"ANNOT_{strat_idx+1:02d}"
        das_data  = X[idx, 0] if X.ndim == 4 else X[idx]
        geo_data  = das_data.mean(axis=0) if das_data.ndim == 2 else das_data
        gt        = int(y[idx])
        p         = int(pred[idx])
        c         = float(conf[idx])
        geo_conf  = c if p == 1 else 1.0 - c

        # Get evidence
        init = pipeline.run(
            sample_id=sample_id, das_data=das_data, fs=SAMPLING_RATE,
            ground_truth=gt, model_prediction=p, model_confidence=c,
            explanation="placeholder", strategy_name="_init",
            geo_data=geo_data, geo_confidence=geo_conf,
        )
        evidence = init.symbolic_evidence
        ev_str   = ev_summary(evidence)

        print(f"\n[{sample_id}] {strategy} — GT={gt} Pred={p} Conf={c:.2f}", end=" ")

        gen_result = gen.generate_full(
            strategy=strategy, prediction=p, confidence=c, evidence=evidence
        )

        if not gen_result.success or not gen_result.explanation:
            print("FAILED")
            explanation_text = "(generation failed)"
        else:
            print(f"done ({gen_result.duration_s:.1f}s)")
            explanation_text = gen_result.explanation

        rows.append({
            "annot_id":        f"{sample_id}_{strategy}",
            "sample_id":       sample_id,
            "strategy":        strategy,
            "ground_truth":    gt,
            "prediction":      p,
            "confidence":      round(c, 3),
            "correct":         int(p == gt),
            "symbolic_evidence": ev_str,
            "explanation_text":  explanation_text,
        })
        call_num += 1

    # Also add 5 more samples with P1_raw for comparison richness
    print("\nGenerating 5 additional P1_raw explanations...")
    for extra_i in range(1, 6):
        idx       = sample_indices[extra_i % len(sample_indices)]
        sample_id = f"ANNOT_P1_{extra_i:02d}"
        das_data  = X[idx, 0] if X.ndim == 4 else X[idx]
        geo_data  = das_data.mean(axis=0) if das_data.ndim == 2 else das_data
        gt, p, c  = int(y[idx]), int(pred[idx]), float(conf[idx])
        geo_conf  = c if p == 1 else 1.0 - c

        init      = pipeline.run(
            sample_id=sample_id, das_data=das_data, fs=SAMPLING_RATE,
            ground_truth=gt, model_prediction=p, model_confidence=c,
            explanation="placeholder", strategy_name="_init",
            geo_data=geo_data, geo_confidence=geo_conf,
        )
        evidence  = init.symbolic_evidence
        ev_str    = ev_summary(evidence)

        print(f"  [{sample_id}] P1_raw", end=" ")
        gen_result = gen.generate_full(
            strategy="P1_raw", prediction=p, confidence=c, evidence=evidence
        )
        if gen_result.success and gen_result.explanation:
            print(f"done ({gen_result.duration_s:.1f}s)")
            explanation_text = gen_result.explanation
        else:
            print("FAILED")
            explanation_text = "(generation failed)"

        rows.append({
            "annot_id":          f"{sample_id}_P1_raw",
            "sample_id":         sample_id,
            "strategy":          "P1_raw",
            "ground_truth":      gt,
            "prediction":        p,
            "confidence":        round(c, 3),
            "correct":           int(p == gt),
            "symbolic_evidence": ev_str,
            "explanation_text":  explanation_text,
        })

    print(f"\nTotal explanations: {len(rows)}")

    # Save raw CSV
    csv_path = os.path.join(RESULTS_DIR, "annotation_samples.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    print(f"Saved: {csv_path}")

    # Build annotation Excel sheets
    try:
        build_annotation_sheets(rows, RESULTS_DIR)
    except Exception as e:
        print(f"Excel generation failed ({e}) — CSV is still available")


def build_annotation_sheets(rows, out_dir):
    """Build two identical blank annotation Excel sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed — run: pip install openpyxl")
        return

    HEADER_FILL  = PatternFill("solid", start_color="1F4E79")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    INST_FILL    = PatternFill("solid", start_color="FFF2CC")
    INST_FONT    = Font(name="Arial", size=9, italic=True)
    SAMPLE_FILL  = PatternFill("solid", start_color="D6E4F0")
    SAMPLE_FONT  = Font(bold=True, name="Arial", size=10)
    EXPL_FILL    = PatternFill("solid", start_color="FAFAFA")
    THIN         = Side(style="thin", color="AAAAAA")
    THIN_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    WRAP         = Alignment(wrap_text=True, vertical="top")

    for annotator in ["A", "B"]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Annotations"

        # ── Instructions row ────────────────────────────────────────────────
        ws.merge_cells("A1:H1")
        ws["A1"] = (
            "SAFE-GEM ANNOTATION STUDY — Annotator " + annotator + "  |  "
            "Instructions: For each explanation, read the text and "
            "identify EVERY signal-related factual claim. "
            "Write one claim per row. For each claim, (1) state what "
            "feature it references, (2) note the exact phrase, "
            "(3) decide if it is SUPPORTED, CONTRADICTED, or "
            "UNAVAILABLE based on the symbolic evidence shown. "
            "Mark N/A if no claims found."
        )
        ws["A1"].font  = INST_FONT
        ws["A1"].fill  = INST_FILL
        ws["A1"].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[1].height = 50

        # ── Column headers ───────────────────────────────────────────────────
        headers = [
            "Annot ID",          # A
            "Strategy",          # B
            "GT | Pred | Conf",  # C
            "Symbolic Evidence", # D
            "Explanation Text",  # E
            "Claim # (1,2,3..)", # F
            "Claim Phrase (exact words)", # G
            "Feature Group\n(SNR/Freq/Onset/Amp/DAS/Event/Uncertainty)", # H
            "Verdict\n(SUPPORTED / CONTRADICTED / UNAVAILABLE)", # I
            "Notes (optional)",  # J
        ]
        widths = [18, 16, 18, 35, 55, 10, 40, 25, 20, 25]
        row = 2
        for col, (h, w) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(wrap_text=True, horizontal="center",
                                       vertical="center")
            cell.border    = THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[row].height = 40
        ws.freeze_panes = "A3"

        row = 3
        for r in rows:
            # Sample header row (spans 3 claim rows)
            ws.merge_cells(f"A{row}:A{row+2}")
            ws.merge_cells(f"B{row}:B{row+2}")
            ws.merge_cells(f"C{row}:C{row+2}")
            ws.merge_cells(f"D{row}:D{row+2}")
            ws.merge_cells(f"E{row}:E{row+2}")

            for col, val in enumerate([
                r["annot_id"],
                r["strategy"],
                f"GT={r['ground_truth']} | Pred={r['prediction']} | "
                f"Conf={r['confidence']} | "
                f"{'CORRECT' if r['correct'] else 'WRONG'}",
                r["symbolic_evidence"],
                r["explanation_text"],
            ], start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill      = SAMPLE_FILL if col < 5 else EXPL_FILL
                cell.font      = SAMPLE_FONT if col < 3 else Font(name="Arial", size=9)
                cell.alignment = WRAP
                cell.border    = THIN_BORDER
                ws.row_dimensions[row].height = 80

            # Claim rows (3 blank rows per explanation)
            for claim_row in range(row, row + 3):
                ws.row_dimensions[claim_row].height = 30
                for col in range(6, 11):   # columns F–J
                    cell = ws.cell(row=claim_row, column=col)
                    cell.border    = THIN_BORDER
                    cell.alignment = WRAP
                    if col == 6:   # Claim #
                        cell.value = claim_row - row + 1
                    if col == 9:   # Verdict dropdown hint
                        cell.fill = PatternFill("solid", start_color="FFF8E1")

            row += 3

        # ── Summary sheet ────────────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "Annotator"
        ws2["B1"] = annotator
        ws2["A2"] = "Date completed"
        ws2["B2"] = ""
        ws2["A3"] = "Total explanations reviewed"
        ws2["B3"] = len(rows)
        ws2["A4"] = "Total claims identified"
        ws2["B4"] = "=COUNTA(Annotations!G3:G999)-COUNTIF(Annotations!G3:G999,\"N/A\")"
        ws2["A5"] = "Supported claims"
        ws2["B5"] = "=COUNTIF(Annotations!I3:I999,\"SUPPORTED\")"
        ws2["A6"] = "Contradicted claims"
        ws2["B6"] = "=COUNTIF(Annotations!I3:I999,\"CONTRADICTED\")"
        ws2["A7"] = "Unavailable claims"
        ws2["B7"] = "=COUNTIF(Annotations!I3:I999,\"UNAVAILABLE\")"

        for r_idx in range(1, 8):
            ws2.cell(r_idx, 1).font = Font(bold=True, name="Arial", size=10)
            ws2.column_dimensions["A"].width = 30
            ws2.column_dimensions["B"].width = 20

        path = os.path.join(out_dir, f"annotation_sheet_{annotator}.xlsx")
        wb.save(path)
        print(f"Saved: annotation_sheet_{annotator}.xlsx")
