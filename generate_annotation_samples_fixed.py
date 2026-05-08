"""
generate_annotation_samples_fixed.py
------------------------------------
Generates the real 30 explanation samples needed for the MC2 annotation study.

What it creates in RESULTS_DIR:
  - annotation_samples.csv      : 30 explanations with symbolic evidence
  - annotation_sheet_A.xlsx     : blank annotation sheet for Annotator A
  - annotation_sheet_B.xlsx     : blank annotation sheet for Annotator B

Run from your safe_gem folder:
    python .\generate_annotation_samples_fixed.py

Main fixes compared with the previous script:
  1. build_annotation_sheets() is defined before it is called.
  2. The generation loop is 6 samples x 5 strategies = 30 explanations.
  3. The extra 5 P1_raw explanations are removed, so the CSV has exactly 30 rows.
  4. Each explanation has 8 blank claim rows in Excel, not only 3.
"""

import csv
import os
import random
import sys
from pathlib import Path

import numpy as np

# Keep imports working when this script is placed in the safe_gem folder.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# You can override these from PowerShell if needed:
#   $env:SAFE_GEM_DATA_DIR="C:\path\to\Dataset"
#   $env:SAFE_GEM_RESULTS_DIR="C:\path\to\Results"
DATA_DIR = os.environ.get(
    "SAFE_GEM_DATA_DIR",
    r"C:\Users\Valued User\Desktop\Office PhD\agentic_egs\agentic_egs\agentic_egs\Dataset",
)
RESULTS_DIR = os.environ.get(
    "SAFE_GEM_RESULTS_DIR",
    r"C:\Users\Valued User\Desktop\Office PhD\agentic_egs\agentic_egs\agentic_egs\Results",
)

SAMPLING_RATE = 500.0
STRATEGIES = [
    "P1_raw",
    "P2_grounded",
    "P3_uncertainty",
    "P4_constrained",
    "P5_safe_gem",
]
N_SAMPLES = 6                 # 6 samples
CLAIM_ROWS_PER_EXPLANATION = 8 # More room for annotators than the old 3 rows
RANDOM_SEED = 42
MODEL_NAME = "qwen2.5:7b"     # Change to "gemma4" etc. if that is what your generator uses.


def select_samples(y, pred, conf, n_samples=6, seed=42):
    """Select a diverse set of samples: clear events, wrong noise, and ambiguous cases."""
    rng = random.Random(seed)
    idx = list(range(len(y)))

    clear_events = [i for i in idx if y[i] == 1 and pred[i] == 1 and conf[i] > 0.95]
    wrong_noise = [i for i in idx if y[i] == 0 and pred[i] == 1 and conf[i] > 0.95]
    ambiguous = sorted(
        [i for i in idx if 0.45 < conf[i] < 0.65],
        key=lambda i: abs(float(conf[i]) - 0.5),
    )

    chosen = clear_events[:2] + wrong_noise[:2] + ambiguous[:2]

    # Fallback: if any category has fewer than expected, fill from remaining diverse samples.
    if len(chosen) < n_samples:
        remaining = [i for i in idx if i not in chosen]
        remaining = sorted(remaining, key=lambda i: abs(float(conf[i]) - 0.5))
        chosen.extend(remaining[: n_samples - len(chosen)])

    # Final guard against duplicates and over-selection.
    deduped = []
    for i in chosen:
        if i not in deduped:
            deduped.append(i)

    if len(deduped) < n_samples:
        extra = [i for i in idx if i not in deduped]
        rng.shuffle(extra)
        deduped.extend(extra[: n_samples - len(deduped)])

    return deduped[:n_samples]


def ev_summary(ev):
    """Short string summary of symbolic evidence."""
    parts = []
    fields = [
        ("SNR", "snr_level"),
        ("Freq", "frequency_band"),
        ("Onset", "onset_character"),
        ("Amp", "amplitude_contrast"),
        ("DAS", "das_coherence"),
        ("EvtSupport", "event_support"),
        ("Ambiguity", "signal_ambiguity"),
    ]
    for label, attr in fields:
        val = getattr(ev, attr, "unavailable")
        if val not in ("unavailable", "", None):
            parts.append(f"{label}:{val}")
    return " | ".join(parts) if parts else "No symbolic evidence available"


def build_annotation_sheets(rows, out_dir):
    """Build two identical blank annotation Excel sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed — run: pip install openpyxl")
        return

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    inst_fill = PatternFill("solid", start_color="FFF2CC")
    inst_font = Font(name="Arial", size=9, italic=True)
    sample_fill = PatternFill("solid", start_color="D6E4F0")
    sample_font = Font(bold=True, name="Arial", size=10)
    expl_fill = PatternFill("solid", start_color="FAFAFA")
    thin = Side(style="thin", color="AAAAAA")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    for annotator in ["A", "B"]:
        wb = Workbook()
        ws = wb.active
        ws.title = "Annotations"

        ws.merge_cells("A1:J1")
        ws["A1"] = (
            f"SAFE-GEM ANNOTATION STUDY — Annotator {annotator} | "
            "Instructions: For each explanation, read the text and identify EVERY "
            "signal-related factual claim. Write one claim per row. For each claim, "
            "copy the exact phrase, choose the feature group, and mark the verdict as "
            "SUPPORTED, CONTRADICTED, or UNAVAILABLE based on the symbolic evidence shown. "
            "Use N/A only if the explanation contains no signal-related factual claim."
        )
        ws["A1"].font = inst_font
        ws["A1"].fill = inst_fill
        ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[1].height = 60

        headers = [
            "Annot ID",
            "Strategy",
            "GT | Pred | Conf",
            "Symbolic Evidence",
            "Explanation Text",
            "Claim #",
            "Claim Phrase (exact words)",
            "Feature Group\n(SNR/Freq/Onset/Amp/DAS/Event/Uncertainty)",
            "Verdict\n(SUPPORTED / CONTRADICTED / UNAVAILABLE)",
            "Notes (optional)",
        ]
        widths = [22, 16, 20, 38, 60, 10, 42, 28, 24, 25]

        for col, (header, width) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[2].height = 42
        ws.freeze_panes = "A3"

        row = 3
        for r in rows:
            end_row = row + CLAIM_ROWS_PER_EXPLANATION - 1
            for col_letter in ["A", "B", "C", "D", "E"]:
                ws.merge_cells(f"{col_letter}{row}:{col_letter}{end_row}")

            fixed_values = [
                r["annot_id"],
                r["strategy"],
                f"GT={r['ground_truth']} | Pred={r['prediction']} | Conf={r['confidence']} | "
                f"{'CORRECT' if r['correct'] else 'WRONG'}",
                r["symbolic_evidence"],
                r["explanation_text"],
            ]
            for col, value in enumerate(fixed_values, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.fill = sample_fill if col < 5 else expl_fill
                cell.font = sample_font if col < 3 else Font(name="Arial", size=9)
                cell.alignment = wrap
                cell.border = thin_border

            for claim_row in range(row, end_row + 1):
                ws.row_dimensions[claim_row].height = 34
                for col in range(6, 11):
                    cell = ws.cell(row=claim_row, column=col)
                    cell.border = thin_border
                    cell.alignment = wrap
                    if col == 6:
                        cell.value = claim_row - row + 1
                    if col == 9:
                        cell.fill = PatternFill("solid", start_color="FFF8E1")

            row = end_row + 1

        ws2 = wb.create_sheet("Summary")
        summary_rows = [
            ("Annotator", annotator),
            ("Date completed", ""),
            ("Total explanations reviewed", len(rows)),
            ("Total claims identified", '=COUNTA(Annotations!G3:G2000)-COUNTIF(Annotations!G3:G2000,"N/A")'),
            ("Supported claims", '=COUNTIF(Annotations!I3:I2000,"SUPPORTED")'),
            ("Contradicted claims", '=COUNTIF(Annotations!I3:I2000,"CONTRADICTED")'),
            ("Unavailable claims", '=COUNTIF(Annotations!I3:I2000,"UNAVAILABLE")'),
        ]
        for r_idx, (label, value) in enumerate(summary_rows, start=1):
            ws2.cell(r_idx, 1, label).font = Font(bold=True, name="Arial", size=10)
            ws2.cell(r_idx, 2, value)
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 20

        path = out_dir / f"annotation_sheet_{annotator}.xlsx"
        wb.save(path)
        print(f"Saved: {path}")


def generate_single_row(sample_number, idx, strategy, X, y, pred, conf, pipeline, gen):
    """Generate one explanation row for one sample-strategy pair."""
    sample_id = f"ANNOT_{sample_number:02d}"
    annot_id = f"{sample_id}_{strategy}"

    das_data = X[idx, 0] if X.ndim == 4 else X[idx]
    geo_data = das_data.mean(axis=0) if getattr(das_data, "ndim", 0) == 2 else das_data
    gt = int(y[idx])
    p = int(pred[idx])
    c = float(conf[idx])
    geo_conf = c if p == 1 else 1.0 - c

    init = pipeline.run(
        sample_id=sample_id,
        das_data=das_data,
        fs=SAMPLING_RATE,
        ground_truth=gt,
        model_prediction=p,
        model_confidence=c,
        explanation="placeholder",
        strategy_name="_init",
        geo_data=geo_data,
        geo_confidence=geo_conf,
    )
    evidence = init.symbolic_evidence
    ev_str = ev_summary(evidence)

    print(f"[{annot_id}] sample_index={idx} GT={gt} Pred={p} Conf={c:.2f}", end=" ")
    gen_result = gen.generate_full(strategy=strategy, prediction=p, confidence=c, evidence=evidence)

    if not gen_result.success or not gen_result.explanation:
        print("FAILED")
        explanation_text = "(generation failed)"
    else:
        print(f"done ({gen_result.duration_s:.1f}s)")
        explanation_text = gen_result.explanation

    return {
        "annot_id": annot_id,
        "sample_id": sample_id,
        "sample_index": int(idx),
        "strategy": strategy,
        "ground_truth": gt,
        "prediction": p,
        "confidence": round(c, 3),
        "correct": int(p == gt),
        "symbolic_evidence": ev_str,
        "explanation_text": explanation_text,
    }


def main():
    from safe_gem import SafeGEMPipeline
    from generate_explanation import ExplanationGenerator

    results_dir = Path(RESULTS_DIR)
    results_dir.mkdir(parents=True, exist_ok=True)

    X = np.load(os.path.join(DATA_DIR, "X_forge.npy"), mmap_mode="r")
    y = np.load(os.path.join(DATA_DIR, "y_forge.npy"))
    pred = np.load(os.path.join(DATA_DIR, "forge_predictions.npy"))
    conf = np.load(os.path.join(DATA_DIR, "forge_confidences.npy"))

    sample_indices = select_samples(y, pred, conf, n_samples=N_SAMPLES, seed=RANDOM_SEED)
    print(f"Selected {len(sample_indices)} samples: {sample_indices}")
    print(f"Generating {len(sample_indices)} x {len(STRATEGIES)} = {len(sample_indices) * len(STRATEGIES)} explanations")

    pipeline = SafeGEMPipeline(pre_event_s=0.5)
    gen = ExplanationGenerator(model=MODEL_NAME)
    gen.warmup()

    rows = []
    for sample_number, idx in enumerate(sample_indices, start=1):
        for strategy in STRATEGIES:
            row = generate_single_row(sample_number, idx, strategy, X, y, pred, conf, pipeline, gen)
            rows.append(row)

    print(f"\nTotal explanations: {len(rows)}")
    if len(rows) != N_SAMPLES * len(STRATEGIES):
        raise RuntimeError(f"Expected {N_SAMPLES * len(STRATEGIES)} rows but got {len(rows)}")

    csv_path = results_dir / "annotation_samples.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    print(f"Saved: {csv_path}")

    build_annotation_sheets(rows, results_dir)
    print("\nDone. Send annotation_sheet_A.xlsx and annotation_sheet_B.xlsx to the two annotators.")


if __name__ == "__main__":
    main()
